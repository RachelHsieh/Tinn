import pandas as pd
import os
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelVendorSplitter:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Vendor Splitter")
        self.root.geometry("600x500")
        self.root.resizable(True, True)
        
        # Variables
        self.input_file = tk.StringVar()
        self.vendor_column = tk.StringVar()
        self.model_year = tk.StringVar(value="MY26")
        self.month = tk.StringVar(value="SEP")
        self.available_columns = []
        self.header_row = 0  # Track which row contains headers
        
        self.setup_ui()
        
    def setup_ui(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="Excel Vendor Splitter", font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # File selection
        ttk.Label(main_frame, text="Excel File:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.input_file, width=50).grid(row=1, column=1, sticky=(tk.W, tk.E), pady=5, padx=(5, 5))
        ttk.Button(main_frame, text="Browse", command=self.browse_file).grid(row=1, column=2, pady=5)
        
        # Vendor column selection
        ttk.Label(main_frame, text="Vendor Column:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.vendor_combo = ttk.Combobox(main_frame, textvariable=self.vendor_column, width=47)
        self.vendor_combo.grid(row=2, column=1, sticky=(tk.W, tk.E), pady=5, padx=(5, 5))
        ttk.Button(main_frame, text="Refresh", command=self.load_columns).grid(row=2, column=2, pady=5)
        
        # Model year input
        ttk.Label(main_frame, text="Model Year:").grid(row=3, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.model_year, width=20).grid(row=3, column=1, sticky=tk.W, pady=5, padx=(5, 5))
        
        # Month input
        ttk.Label(main_frame, text="Month:").grid(row=4, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.month, width=20).grid(row=4, column=1, sticky=tk.W, pady=5, padx=(5, 5))
        
        # Process button
        ttk.Button(main_frame, text="Split Excel File", command=self.process_file, style="Accent.TButton").grid(
            row=5, column=0, columnspan=3, pady=20, sticky=(tk.W, tk.E))
        
        # Progress bar
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=6, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        # Status text area
        ttk.Label(main_frame, text="Status:").grid(row=7, column=0, sticky=tk.W, pady=(10, 5))
        
        # Text widget with scrollbar
        text_frame = ttk.Frame(main_frame)
        text_frame.grid(row=8, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)
        
        self.status_text = tk.Text(text_frame, height=12, wrap=tk.WORD)
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.status_text.yview)
        self.status_text.configure(yscrollcommand=scrollbar.set)
        
        self.status_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Configure main_frame row weights
        main_frame.rowconfigure(8, weight=1)
        
    def browse_file(self):
        filename = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*")
            ]
        )
        if filename:
            self.input_file.set(filename)
            self.load_columns()
            
    def load_columns(self):
        if not self.input_file.get():
            return
            
        try:
            self.log_status("Loading column names from 'PO list' sheet...")
            
            # First, try to detect the header row by reading a few rows from "PO list" sheet
            df_preview = pd.read_excel(self.input_file.get(), sheet_name="PO list", nrows=5, header=None)
            
            # Look for the row that contains column headers
            header_row = 0
            for i in range(min(5, len(df_preview))):
                row_values = df_preview.iloc[i].astype(str).tolist()
                # Check if this row contains likely header names
                if any(keyword in str(val).lower() for val in row_values 
                      for keyword in ['vendor', 'po', 'number', 'model', 'category', 'item', 'description']):
                    header_row = i
                    self.log_status(f"Found headers in row {i + 1} of 'PO list' sheet")
                    break
            
            # Read the file with the correct header row from "PO list" sheet
            df = pd.read_excel(self.input_file.get(), sheet_name="PO list", header=header_row, nrows=0)
            
            # Convert all column names to strings and clean them
            self.available_columns = []
            for col in df.columns:
                col_str = str(col).strip()
                # Skip completely unnamed columns
                if not col_str.startswith('Unnamed:') and col_str != 'nan':
                    self.available_columns.append(col_str)
            
            # If we still have mostly unnamed columns, try different header rows
            if len([col for col in self.available_columns if not col.startswith('Unnamed:')]) < 3:
                self.log_status("Trying to find headers in different rows of 'PO list' sheet...")
                for try_header in range(1, 5):
                    try:
                        df_try = pd.read_excel(self.input_file.get(), sheet_name="PO list", header=try_header, nrows=0)
                        potential_columns = [str(col).strip() for col in df_try.columns]
                        named_cols = [col for col in potential_columns if not col.startswith('Unnamed:') and col != 'nan']
                        
                        if len(named_cols) > len(self.available_columns):
                            self.available_columns = potential_columns
                            header_row = try_header
                            self.log_status(f"Better headers found in row {try_header + 1} of 'PO list' sheet")
                    except:
                        continue
            
            # Store the header row for later use
            self.header_row = header_row
            
            self.vendor_combo['values'] = self.available_columns
            
            # Try to auto-select vendor column
            vendor_candidates = []
            for col in self.available_columns:
                col_lower = str(col).lower()
                if 'vendor' in col_lower or 'supplier' in col_lower:
                    vendor_candidates.append(col)
            
            if vendor_candidates:
                self.vendor_column.set(vendor_candidates[0])
            
            self.log_status(f"Found {len(self.available_columns)} columns in 'PO list' sheet: {', '.join(map(str, self.available_columns))}")
            
        except Exception as e:
            self.log_status(f"Error loading columns from 'PO list' sheet: {str(e)}")
            messagebox.showerror("Error", f"Could not read 'PO list' sheet from Excel file: {str(e)}")
    
    def log_status(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.status_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.status_text.see(tk.END)
        self.root.update_idletasks()
    
    def process_file(self):
        # Validate inputs
        if not self.input_file.get():
            messagebox.showerror("Error", "Please select an Excel file")
            return
            
        if not self.vendor_column.get():
            messagebox.showerror("Error", "Please select a vendor column")
            return
            
        if not self.model_year.get():
            messagebox.showerror("Error", "Please enter a model year")
            return
            
        if not self.month.get():
            messagebox.showerror("Error", "Please enter a month")
            return
        
        # Start processing
        self.progress.start(10)
        self.status_text.delete(1.0, tk.END)
        self.log_status("Starting Excel file processing...")
        
        try:
            self.split_excel_by_vendor()
        except Exception as e:
            self.log_status(f"Error: {str(e)}")
            messagebox.showerror("Error", f"Processing failed: {str(e)}")
        finally:
            self.progress.stop()
    
    def split_excel_by_vendor(self):
        input_file = self.input_file.get()
        vendor_column = self.vendor_column.get()
        model_year = self.model_year.get()
        month = self.month.get()
        
        # Define required columns for data cleaning
        required_columns = [
            "Vendor name", "Sbc Market Code", "Po number", "Rider Experience", 
            "Category", "Model", "Item number", "Item description", 
            "Pod Quantity Ordered", "Ship To Loc Name", "prod month", 
            "Upc", "Mpl Model Year", "Need by date"
        ]
        
        # Read the Excel file using the detected header row from "PO list" sheet
        self.log_status(f"Reading 'PO list' sheet from: {os.path.basename(input_file)}")
        header_row = getattr(self, 'header_row', 0)  # Default to 0 if not set
        df = pd.read_excel(input_file, sheet_name="PO list", header=header_row)
        
        # Convert all column names to strings to handle numeric columns
        df.columns = [str(col).strip() for col in df.columns]
        
        # Check if vendor column exists
        vendor_column_str = str(vendor_column)
        if vendor_column_str not in df.columns:
            raise ValueError(f"Column '{vendor_column_str}' not found in the Excel file.")
        
        # Data cleaning - check which required columns exist
        self.log_status("Performing data cleaning...")
        existing_columns = []
        missing_columns = []
        
        # Convert required columns to strings for comparison
        required_columns_str = [str(col) for col in required_columns]
        
        for col in required_columns_str:
            if col in df.columns:
                existing_columns.append(col)
            else:
                missing_columns.append(col)
        
        # Log column status
        if missing_columns:
            self.log_status(f"Warning: Missing columns: {', '.join(missing_columns)}")
        self.log_status(f"Keeping {len(existing_columns)} columns: {', '.join(existing_columns)}")
        
        # Filter dataframe to keep only existing required columns
        df_cleaned = df[existing_columns].copy()
        
        # Get today's date in the required format
        today = datetime.now().strftime("%d-%m-%Y")
        
        # Get unique vendors using string column name
        vendors = df_cleaned[vendor_column_str].dropna().unique()
        self.log_status(f"Found {len(vendors)} unique vendors")
        
        # Create output directory
        output_dir = "vendor_splits"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # Split by vendor and save
        files_created = []
        for i, vendor in enumerate(vendors):
            # Filter data for current vendor using string column name
            vendor_data = df_cleaned[df_cleaned[vendor_column_str] == vendor]
            
            # Clean vendor name for filename
            clean_vendor = str(vendor).replace("/", "_").replace("\\", "_").replace(":", "_").replace("*", "_").replace("?", "_").replace("\"", "_").replace("<", "_").replace(">", "_").replace("|", "_")
            
            # Create filename
            filename = f"TRENDPOWER({clean_vendor})_{model_year} {month}_market PO_{today}.xlsx"
            filepath = os.path.join(output_dir, filename)
            
            # Save to Excel with formatting
            self.save_formatted_excel(vendor_data, filepath)
            files_created.append(filename)
            
            self.log_status(f"Created: {filename} ({len(vendor_data)} rows, {len(vendor_data.columns)} columns)")
            
            # Update progress
            progress_percent = ((i + 1) / len(vendors)) * 100
            self.root.update_idletasks()
        
        self.log_status(f"\nProcess completed successfully!")
        self.log_status(f"Created {len(files_created)} files in '{output_dir}' directory")
        self.log_status(f"Each file contains only the {len(existing_columns)} required columns")
        self.log_status("Files formatted with Calibri Light font, blue headers (#D9E1F2), Short Date format, and UPC as Fraction")
        
        # Show completion message
        messagebox.showinfo("Success", 
                          f"Successfully created {len(files_created)} vendor files!\n"
                          f"Files saved in: {os.path.abspath(output_dir)}\n"
                          f"Data cleaned to keep only {len(existing_columns)} required columns\n"
                          f"Formatted with Calibri Light font, blue headers, Short Date format, and UPC as Fraction")

    def save_formatted_excel(self, df, filepath):
        """
        Save DataFrame to Excel with Calibri Light font and Short Date formatting
        """
        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            
            # Add data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Create Calibri Light font style
            calibri_light = Font(name='Calibri Light', size=11)
            
            # Create header background fill
            header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            # Create Short Date style
            date_style = NamedStyle(name="short_date")
            date_style.font = calibri_light
            date_style.number_format = 'M/D/YYYY'
            
            # Apply formatting to all cells
            for row in ws.iter_rows():
                for cell in row:
                    # Apply Calibri Light font to all cells
                    cell.font = calibri_light
                    
                    # Get header name for this column
                    header_cell = ws.cell(row=1, column=cell.column)
                    header_name = ""
                    if header_cell.value and isinstance(header_cell.value, str):
                        header_name = str(header_cell.value).lower().strip()
                    
                    # Check if cell contains date-like data
                    if cell.value and isinstance(cell.value, str):
                        # Apply date formatting for date columns
                        if any(date_keyword in header_name for date_keyword in 
                              ['date', 'month', 'year', 'time']):
                            # Try to parse as date
                            try:
                                if '/' in str(cell.value) or '-' in str(cell.value):
                                    cell.number_format = 'M/D/YYYY'
                            except:
                                pass
                    
                    # Handle actual datetime objects
                    elif isinstance(cell.value, datetime):
                        cell.number_format = 'M/D/YYYY'
                    
                    # Format UPC column as Fraction
                    if header_name == 'upc' and cell.value is not None and cell.row > 1:
                        try:
                            # Convert to number if it's a string
                            if isinstance(cell.value, str):
                                # Remove any non-numeric characters except decimal point
                                numeric_value = ''.join(c for c in cell.value if c.isdigit() or c == '.')
                                if numeric_value:
                                    cell.value = float(numeric_value)
                            
                            # Apply fraction format
                            cell.number_format = '# ?/?'
                        except:
                            # If conversion fails, leave as is
                            pass
            
            # Make header row bold with background color
            for cell in ws[1]:
                if cell.value:
                    cell.font = Font(name='Calibri Light', size=11, bold=True)
                    cell.fill = header_fill
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook
            wb.save(filepath)
            
        except Exception as e:
            # Fallback to pandas if openpyxl formatting fails
            self.log_status(f"Warning: Could not apply formatting, saving as standard Excel: {str(e)}")
            df.to_excel(filepath, index=False)

def main():
    root = tk.Tk()
    app = ExcelVendorSplitter(root)
    root.mainloop()

if __name__ == "__main__":
    main()